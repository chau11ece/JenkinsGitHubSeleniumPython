import openpyxl

book = openpyxl.load_workbook("C:\\Users\\DELL\\Downloads\\excel_data.xlsx")
sheet = book.active
cell_B1 = sheet.cell(row=1, column=2)
print(cell_B1.value)
cell_B2 = sheet.cell(row=2, column=2)
cell_B2.value = "Chau"
print(cell_B2.value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row+1):
    # run tc2 only
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
